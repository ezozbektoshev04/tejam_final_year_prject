import json
from datetime import datetime, timedelta
from flask import Blueprint, jsonify, request
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db, User, Shop, Order, FoodItem, PlatformSetting, Notification, ShopPayout
from utils.email import send_shop_approved_email, send_account_deleted_email, send_shop_status_email

admin_bp = Blueprint("admin", __name__)


def require_admin():
    """Returns current user if admin, else None."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or user.role != "admin":
        return None
    return user


# ── Stats overview ────────────────────────────────────────────────────────────

@admin_bp.route("/stats", methods=["GET"])
@jwt_required()
def stats():
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    total_customers = User.query.filter_by(role="customer").count()
    total_shops = User.query.filter_by(role="shop", is_approved=True).count()
    pending_shop_approvals = User.query.filter_by(role="shop", is_approved=False).count()
    total_orders = Order.query.count()
    total_listings = FoodItem.query.count()
    revenue = db.session.query(db.func.sum(Order.total_price)).filter(
        Order.status == "picked_up"
    ).scalar() or 0

    pending_orders = Order.query.filter_by(status="pending").count()
    confirmed_orders = Order.query.filter_by(status="confirmed").count()

    return jsonify({
        "total_customers": total_customers,
        "total_shops": total_shops,
        "pending_shop_approvals": pending_shop_approvals,
        "total_orders": total_orders,
        "total_listings": total_listings,
        "total_revenue": round(revenue),
        "pending_orders": pending_orders,
        "confirmed_orders": confirmed_orders,
    })


# ── Users ─────────────────────────────────────────────────────────────────────

@admin_bp.route("/users", methods=["GET"])
@jwt_required()
def list_users():
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    role_filter = request.args.get("role")  # 'customer' | 'shop'
    query = User.query.filter(User.role != "admin")
    if role_filter:
        query = query.filter_by(role=role_filter)

    users = query.order_by(User.created_at.desc()).all()

    result = []
    for u in users:
        data = u.to_dict()
        if u.role == "customer":
            data["order_count"] = Order.query.filter_by(customer_id=u.id).count()
        if u.role == "shop" and u.shops:
            shop_ids = [s.id for s in u.shops]
            data["shops"] = [s.to_dict() for s in u.shops]
            data["listing_count"] = FoodItem.query.filter(FoodItem.shop_id.in_(shop_ids)).count()
            data["order_count"] = Order.query.join(FoodItem).filter(
                FoodItem.shop_id.in_(shop_ids)
            ).count()
        result.append(data)

    return jsonify(result)


@admin_bp.route("/users/<int:user_id>", methods=["DELETE"])
@jwt_required()
def delete_user(user_id):
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    user = User.query.get_or_404(user_id)
    if user.role == "admin":
        return jsonify({"error": "Cannot delete admin account"}), 400

    email, name, role = user.email, user.name, user.role
    db.session.delete(user)
    db.session.commit()
    send_account_deleted_email(email, name, role)
    return jsonify({"message": "User deleted"})


# ── Shops ─────────────────────────────────────────────────────────────────────

@admin_bp.route("/shops", methods=["GET"])
@jwt_required()
def list_shops():
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    shops = Shop.query.order_by(Shop.id).all()
    result = []
    for s in shops:
        data = s.to_dict()
        owner = User.query.get(s.user_id)
        data["owner_name"] = owner.name if owner else None
        data["owner_email"] = owner.email if owner else None
        data["listing_count"] = FoodItem.query.filter_by(shop_id=s.id).count()
        data["order_count"] = Order.query.join(FoodItem).filter(
            FoodItem.shop_id == s.id
        ).count()
        result.append(data)
    return jsonify(result)


@admin_bp.route("/shops/<int:shop_id>/toggle", methods=["PUT"])
@jwt_required()
def toggle_shop(shop_id):
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    shop = Shop.query.get_or_404(shop_id)
    shop.is_active = not shop.is_active
    db.session.commit()

    owner = User.query.get(shop.user_id)
    if owner:
        send_shop_status_email(owner.email, owner.name, shop.name, deactivated=not shop.is_active)

    return jsonify({"id": shop.id, "is_active": shop.is_active})


# ── Platform Settings ─────────────────────────────────────────────────────────

ALLOWED_KEYS = {
    "categories",
    "min_discount_percent",
    "max_discount_percent",
    "low_stock_threshold",
    "commission_rate",
    "notification_order_placed",
    "notification_order_confirmed",
    "notification_order_picked_up",
    "notification_order_cancelled",
}


@admin_bp.route("/settings", methods=["GET"])
@jwt_required()
def get_settings():
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    rows = PlatformSetting.query.filter(PlatformSetting.key.in_(ALLOWED_KEYS)).all()
    result = {}
    for row in rows:
        try:
            result[row.key] = json.loads(row.value)
        except Exception:
            result[row.key] = row.value
    return jsonify(result)


@admin_bp.route("/pending-shops", methods=["GET"])
@jwt_required()
def pending_shops():
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    users = User.query.filter_by(role="shop", is_approved=False).order_by(User.created_at.desc()).all()
    result = []
    for u in users:
        data = u.to_dict()
        data["shops"] = [s.to_dict() for s in u.shops]
        result.append(data)
    return jsonify(result)


@admin_bp.route("/approve-shop/<int:user_id>", methods=["POST"])
@jwt_required()
def approve_shop(user_id):
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    user = User.query.get_or_404(user_id)
    if user.role != "shop":
        return jsonify({"error": "User is not a shop owner"}), 400

    user.is_approved = True
    shop_name = user.shops[0].name if user.shops else user.name
    notif = Notification(
        user_id=user.id,
        message="Your shop has been approved! You can now log in and start listing your surplus food.",
        link="/dashboard",
    )
    db.session.add(notif)
    db.session.commit()
    send_shop_approved_email(user.email, user.name, shop_name)
    return jsonify({"message": f"Shop owner '{user.name}' approved successfully."})


@admin_bp.route("/reject-shop/<int:user_id>", methods=["DELETE"])
@jwt_required()
def reject_shop(user_id):
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    user = User.query.get_or_404(user_id)
    if user.role != "shop":
        return jsonify({"error": "User is not a shop owner"}), 400
    if user.is_approved:
        return jsonify({"error": "Cannot reject an already approved shop"}), 400

    email, name = user.email, user.name
    db.session.delete(user)
    db.session.commit()
    send_account_deleted_email(email, name, "shop")
    return jsonify({"message": "Shop application rejected and account removed."})


@admin_bp.route("/earnings", methods=["GET"])
@jwt_required()
def earnings():
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    from collections import defaultdict

    # ── Filters ──
    start_str = request.args.get("start")
    end_str   = request.args.get("end")
    search    = request.args.get("search", "").strip().lower()
    page      = request.args.get("page", 1, type=int)
    per_page  = request.args.get("per_page", 10, type=int)

    query = Order.query.filter_by(status="picked_up")
    if start_str:
        try:
            query = query.filter(Order.created_at >= datetime.strptime(start_str, "%Y-%m-%d"))
        except ValueError:
            pass
    if end_str:
        try:
            end_dt = datetime.strptime(end_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query = query.filter(Order.created_at <= end_dt)
        except ValueError:
            pass

    completed = query.all()

    total_commission = sum(o.commission_amount or 0 for o in completed)
    total_revenue    = sum(o.total_price for o in completed)
    total_payout     = sum(o.shop_payout or 0 for o in completed)

    # ── This month / last month (always all-time, not affected by filter) ──
    now = datetime.utcnow()
    month_start      = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_start = (month_start - timedelta(days=1)).replace(day=1)
    all_completed    = Order.query.filter_by(status="picked_up").all()
    this_month_commission = sum(o.commission_amount or 0 for o in all_completed if o.created_at >= month_start)
    last_month_commission = sum(o.commission_amount or 0 for o in all_completed if last_month_start <= o.created_at < month_start)

    # ── Monthly trend (based on filtered data) ──
    monthly_map = defaultdict(lambda: {"revenue": 0.0, "commission": 0.0, "orders": 0})
    for o in completed:
        key = o.created_at.strftime("%Y-%m")
        monthly_map[key]["revenue"]    += o.total_price
        monthly_map[key]["commission"] += o.commission_amount or 0
        monthly_map[key]["orders"]     += 1

    monthly_trend = [
        {"month": k, "revenue": round(v["revenue"]), "commission": round(v["commission"]), "orders": v["orders"]}
        for k, v in sorted(monthly_map.items())
    ]

    # ── Per-shop breakdown ──
    shop_stats = defaultdict(lambda: {
        "orders": 0, "revenue": 0.0, "commission": 0.0, "payout": 0.0,
        "shop_name": "", "city": "", "category": ""
    })
    for o in completed:
        item = o.food_item
        if item and item.shop:
            sid = item.shop.id
            shop_stats[sid]["shop_id"]    = sid
            shop_stats[sid]["shop_name"]  = item.shop.name
            shop_stats[sid]["city"]       = item.shop.city or ""
            shop_stats[sid]["category"]   = item.shop.category or ""
            shop_stats[sid]["orders"]     += 1
            shop_stats[sid]["revenue"]    += o.total_price
            shop_stats[sid]["commission"] += o.commission_amount or 0
            shop_stats[sid]["payout"]     += o.shop_payout or 0

    per_shop_list = sorted(shop_stats.values(), key=lambda x: x["commission"], reverse=True)

    # Attach settlement info (all-time, not period-filtered)
    for s in per_shop_list:
        s["revenue"]    = round(s["revenue"])
        s["commission"] = round(s["commission"])
        s["payout"]     = round(s["payout"])
        settled = db.session.query(db.func.sum(ShopPayout.amount)).filter_by(
            shop_id=s["shop_id"]
        ).scalar() or 0
        all_commission = db.session.query(db.func.sum(Order.commission_amount)).join(FoodItem).filter(
            FoodItem.shop_id == s["shop_id"],
            Order.status == "picked_up"
        ).scalar() or 0
        s["total_settled"]  = round(settled)
        s["pending_payout"] = round(max(0.0, all_commission - settled))

    # ── Search ──
    if search:
        per_shop_list = [
            s for s in per_shop_list
            if search in s["shop_name"].lower()
            or search in s["city"].lower()
            or search in s["category"].lower()
        ]

    # ── Pagination ──
    total_shops    = len(per_shop_list)
    per_shop_page  = per_shop_list[(page - 1) * per_page : page * per_page]

    commission_rate = PlatformSetting.get("commission_rate", 0.10)

    return jsonify({
        "total_commission":      round(total_commission),
        "total_revenue":         round(total_revenue),
        "total_payout":          round(total_payout),
        "this_month_commission": round(this_month_commission),
        "last_month_commission": round(last_month_commission),
        "completed_orders":      len(completed),
        "commission_rate":       round(float(commission_rate), 4),
        "monthly_trend":         monthly_trend,
        "per_shop":              per_shop_page,
        "total_shops":           total_shops,
        "page":                  page,
        "per_page":              per_page,
        "pages":                 max(1, -(-total_shops // per_page)),
    })


@admin_bp.route("/earnings/settle", methods=["POST"])
@jwt_required()
def settle_shop():
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    data    = request.get_json()
    shop_id = data.get("shop_id")
    amount  = data.get("amount")
    note    = data.get("note", "")

    if not shop_id or amount is None:
        return jsonify({"error": "shop_id and amount are required"}), 400

    payout = ShopPayout(
        shop_id    = shop_id,
        amount     = round(float(amount), 2),
        note       = note or None,
        status     = "settled",
        settled_at = datetime.utcnow(),
    )
    db.session.add(payout)
    db.session.commit()
    return jsonify(payout.to_dict())


@admin_bp.route("/earnings/export", methods=["GET"])
@jwt_required()
def export_earnings():
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl is not installed"}), 500

    import io
    from collections import defaultdict

    start_str = request.args.get("start")
    end_str   = request.args.get("end")

    query = Order.query.filter_by(status="picked_up")
    if start_str:
        try:
            query = query.filter(Order.created_at >= datetime.strptime(start_str, "%Y-%m-%d"))
        except ValueError:
            pass
    if end_str:
        try:
            end_dt = datetime.strptime(end_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query = query.filter(Order.created_at <= end_dt)
        except ValueError:
            pass

    completed = query.all()

    total_commission = sum(o.commission_amount or 0 for o in completed)
    total_revenue    = sum(o.total_price for o in completed)
    total_payout     = sum(o.shop_payout or 0 for o in completed)
    commission_rate  = PlatformSetting.get("commission_rate", 0.10)

    period_label = f"{start_str or 'All time'} → {end_str or 'now'}"

    # Monthly trend
    monthly_map = defaultdict(lambda: {"revenue": 0.0, "commission": 0.0, "orders": 0})
    for o in completed:
        key = o.created_at.strftime("%Y-%m")
        monthly_map[key]["revenue"]    += o.total_price
        monthly_map[key]["commission"] += o.commission_amount or 0
        monthly_map[key]["orders"]     += 1

    # Per-shop
    shop_stats = defaultdict(lambda: {
        "orders": 0, "revenue": 0.0, "commission": 0.0, "payout": 0.0,
        "shop_name": "", "city": "", "category": ""
    })
    for o in completed:
        item = o.food_item
        if item and item.shop:
            sid = item.shop.id
            shop_stats[sid]["shop_id"]    = sid
            shop_stats[sid]["shop_name"]  = item.shop.name
            shop_stats[sid]["city"]       = item.shop.city or ""
            shop_stats[sid]["category"]   = item.shop.category or ""
            shop_stats[sid]["orders"]     += 1
            shop_stats[sid]["revenue"]    += o.total_price
            shop_stats[sid]["commission"] += o.commission_amount or 0
            shop_stats[sid]["payout"]     += o.shop_payout or 0

    per_shop_list = sorted(shop_stats.values(), key=lambda x: x["commission"], reverse=True)
    for s in per_shop_list:
        settled = db.session.query(db.func.sum(ShopPayout.amount)).filter_by(shop_id=s["shop_id"]).scalar() or 0
        all_comm = db.session.query(db.func.sum(Order.commission_amount)).join(FoodItem).filter(
            FoodItem.shop_id == s["shop_id"], Order.status == "picked_up"
        ).scalar() or 0
        s["total_settled"]  = round(settled)
        s["pending_payout"] = round(max(0.0, all_comm - settled))

    # ── Build workbook ──
    HEADER_FILL = PatternFill("solid", fgColor="1a7548")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    CENTER      = Alignment(horizontal="center")

    def style_header(ws, row, ncols):
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER

    def auto_width(ws):
        for col in ws.columns:
            w = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)

    wb = openpyxl.Workbook()

    # Sheet 1 — Summary
    ws1 = wb.active
    ws1.title = "Summary"
    rows = [
        ("Tejam · Earnings Report", ""),
        ("", ""),
        ("Period", period_label),
        ("Commission rate", f"{round(float(commission_rate) * 100, 1)}%"),
        ("", ""),
        ("Completed orders", len(completed)),
        ("Total gross revenue (UZS)", round(total_revenue)),
        ("Platform commission earned (UZS)", round(total_commission)),
        ("Total shop payouts (UZS)", round(total_payout)),
    ]
    for i, (label, value) in enumerate(rows, start=1):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=value)
        if i == 1:
            ws1.cell(row=i, column=1).font = Font(bold=True, size=14, color="1a7548")
        elif label:
            ws1.cell(row=i, column=1).font = Font(bold=True, size=10)
        if label in ("Platform commission earned (UZS)",):
            ws1.cell(row=i, column=2).font = Font(bold=True, color="1a7548")
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 30

    # Sheet 2 — Monthly Trend
    ws2 = wb.create_sheet("Monthly Trend")
    headers2 = ["Month", "Orders", "Gross Revenue (UZS)", "Commission (UZS)", "Shop Payout (UZS)"]
    for col, h in enumerate(headers2, start=1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, len(headers2))
    for i, (month, v) in enumerate(sorted(monthly_map.items()), start=2):
        ws2.cell(row=i, column=1, value=month)
        ws2.cell(row=i, column=2, value=v["orders"])
        ws2.cell(row=i, column=3, value=round(v["revenue"]))
        ws2.cell(row=i, column=4, value=round(v["commission"]))
        ws2.cell(row=i, column=5, value=round(v["revenue"] - v["commission"]))
    auto_width(ws2)

    # Sheet 3 — Per Shop
    ws3 = wb.create_sheet("Per Shop")
    headers3 = ["Shop", "City", "Category", "Orders", "Gross Revenue (UZS)",
                 "Commission (UZS)", "Shop Payout (UZS)", "Total Settled (UZS)", "Pending (UZS)"]
    for col, h in enumerate(headers3, start=1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, 1, len(headers3))
    for i, s in enumerate(per_shop_list, start=2):
        ws3.cell(row=i, column=1, value=s["shop_name"])
        ws3.cell(row=i, column=2, value=s["city"])
        ws3.cell(row=i, column=3, value=s["category"])
        ws3.cell(row=i, column=4, value=s["orders"])
        ws3.cell(row=i, column=5, value=round(s["revenue"]))
        ws3.cell(row=i, column=6, value=round(s["commission"]))
        ws3.cell(row=i, column=7, value=round(s["payout"]))
        ws3.cell(row=i, column=8, value=s["total_settled"])
        ws3.cell(row=i, column=9, value=s["pending_payout"])
    auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    filename = f"tejam-earnings-{(start_str or 'all').replace('-', '')}-{(end_str or 'now').replace('-', '')}.xlsx"
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@admin_bp.route("/settings", methods=["PUT"])
@jwt_required()
def update_settings():
    if not require_admin():
        return jsonify({"error": "Admin access required"}), 403

    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    for key, value in data.items():
        if key not in ALLOWED_KEYS:
            continue
        PlatformSetting.set(key, value)

    db.session.commit()
    return jsonify({"message": "Settings updated successfully"})
