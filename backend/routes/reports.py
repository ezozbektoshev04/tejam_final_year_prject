import io
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db, User, Shop, FoodItem, Order

reports_bp = Blueprint("reports", __name__)


def _parse_dates(start_str, end_str):
    """Parse start/end query params, return (start_dt, end_dt) as datetime objects."""
    try:
        start = datetime.strptime(start_str, "%Y-%m-%d")
        end = datetime.strptime(end_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except (ValueError, TypeError):
        return None, None
    return start, end


def _bucket_key(dt, granularity):
    if granularity == "hourly":
        return dt.strftime("%Y-%m-%d %H:00")
    elif granularity == "weekly":
        year, week, _ = dt.isocalendar()
        return f"{year}-W{week:02d}"
    else:  # daily (default)
        return dt.strftime("%Y-%m-%d")


@reports_bp.route("/export", methods=["GET"])
@jwt_required()
def export_report():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl is not installed. Run: pip install openpyxl"}), 500

    user_id = int(get_jwt_identity())
    user = User.query.get_or_404(user_id)

    if user.role != "shop":
        return jsonify({"error": "Shop account required"}), 403

    # --- Resolve which shops to include ---
    shop_id_filter = request.args.get("shop_id", type=int)
    if shop_id_filter:
        owner_ids = [s.id for s in user.shops]
        if shop_id_filter not in owner_ids:
            return jsonify({"error": "Unauthorized"}), 403
        target_shops = [Shop.query.get(shop_id_filter)]
    else:
        target_shops = user.shops

    shop_label = (
        target_shops[0].address.split(",")[0] if len(target_shops) == 1
        else "All branches"
    )

    # --- Date range ---
    start_str = request.args.get("start")
    end_str = request.args.get("end")
    granularity = request.args.get("granularity", "daily")  # hourly | daily | weekly

    if start_str and end_str:
        start_dt, end_dt = _parse_dates(start_str, end_str)
        if not start_dt:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    else:
        # Default: last 30 days
        end_dt = datetime.utcnow()
        start_dt = end_dt - timedelta(days=30)

    period_label = f"{start_dt.strftime('%Y-%m-%d')} → {end_dt.strftime('%Y-%m-%d')}"

    # --- Fetch orders ---
    food_ids = [item.id for s in target_shops for item in s.food_items]
    if food_ids:
        orders = (
            Order.query
            .filter(Order.food_item_id.in_(food_ids))
            .filter(Order.created_at >= start_dt)
            .filter(Order.created_at <= end_dt)
            .order_by(Order.created_at.asc())
            .all()
        )
    else:
        orders = []

    completed = [o for o in orders if o.status in ("confirmed", "picked_up")]
    cancelled = [o for o in orders if o.status == "cancelled"]

    total_revenue     = sum(o.total_price for o in completed)
    total_commission  = sum(o.commission_amount or 0 for o in completed)
    total_payout      = sum(o.shop_payout or 0 for o in completed)
    avg_order_value   = total_revenue / len(completed) if completed else 0

    top_item_name = "—"
    if completed:
        from collections import Counter
        counts = Counter(o.food_item.name for o in completed if o.food_item)
        top_item_name = counts.most_common(1)[0][0] if counts else "—"

    # --- Revenue buckets ---
    from collections import defaultdict
    bucket_orders = defaultdict(int)
    bucket_revenue = defaultdict(float)
    for o in completed:
        key = _bucket_key(o.created_at, granularity)
        bucket_orders[key] += 1
        bucket_revenue[key] += o.total_price

    # Fill in all buckets within range (daily / weekly only; hourly only fills used hours)
    all_buckets = sorted(set(list(bucket_orders.keys()) + list(bucket_revenue.keys())))

    # ========================
    # Build Excel workbook
    # ========================
    wb = openpyxl.Workbook()

    # ---- Styles ----
    HEADER_FILL = PatternFill("solid", fgColor="1a7548")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUBHEADER_FONT = Font(bold=True, size=10)
    CENTER = Alignment(horizontal="center")

    def style_header_row(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # =====================
    # Sheet 1 — Summary
    # =====================
    ws1 = wb.active
    ws1.title = "Summary"

    rows = [
        ("Tejam · Sales Report", ""),
        ("", ""),
        ("Branch", shop_label),
        ("Period", period_label),
        ("Granularity", granularity.capitalize()),
        ("", ""),
        ("Total orders", len(orders)),
        ("Completed orders", len(completed)),
        ("Cancelled orders", len(cancelled)),
        ("", ""),
        ("Total revenue (UZS)", total_revenue),
        ("Platform commission (UZS)", round(total_commission)),
        ("Shop payout (UZS)", round(total_payout)),
        ("Avg order value (UZS)", round(avg_order_value)),
        ("Top-selling item", top_item_name),
    ]

    for i, (label, value) in enumerate(rows, start=1):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=value)
        if i == 1:
            ws1.cell(row=i, column=1).font = Font(bold=True, size=14, color="1a7548")
        elif label in ("Total revenue (UZS)", "Platform commission (UZS)", "Shop payout (UZS)", "Avg order value (UZS)"):
            ws1.cell(row=i, column=2).font = Font(bold=True, color="1a7548")
        elif label and value:
            ws1.cell(row=i, column=1).font = SUBHEADER_FONT

    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 40

    # =====================
    # Sheet 2 — Orders
    # =====================
    ws2 = wb.create_sheet("Orders")
    headers2 = ["Order ID", "Date & Time", "Branch", "Item", "Qty",
                 "Unit Price (UZS)", "Total (UZS)", "Commission (%)", "Commission (UZS)",
                 "Shop Payout (UZS)", "Status", "Payment"]
    for col, h in enumerate(headers2, start=1):
        ws2.cell(row=1, column=col, value=h)
    style_header_row(ws2, 1, len(headers2))

    for row_i, o in enumerate(orders, start=2):
        item = o.food_item
        shop = item.shop if item else None
        branch_label = shop.address.split(",")[0] if shop and shop.address else (shop.name if shop else "—")
        unit_price = round(o.total_price / o.quantity) if o.quantity else o.total_price
        ws2.cell(row=row_i, column=1, value=o.id)
        ws2.cell(row=row_i, column=2, value=o.created_at.strftime("%Y-%m-%d %H:%M"))
        ws2.cell(row=row_i, column=3, value=branch_label)
        ws2.cell(row=row_i, column=4, value=item.name if item else "—")
        ws2.cell(row=row_i, column=5, value=o.quantity)
        ws2.cell(row=row_i, column=6, value=unit_price)
        ws2.cell(row=row_i, column=7,  value=round(o.total_price))
        ws2.cell(row=row_i, column=8,  value=round((o.commission_rate or 0) * 100, 1))
        ws2.cell(row=row_i, column=9,  value=round(o.commission_amount or 0))
        ws2.cell(row=row_i, column=10, value=round(o.shop_payout or 0))
        ws2.cell(row=row_i, column=11, value=o.status)
        ws2.cell(row=row_i, column=12, value=o.payment_method or "cash")

    auto_width(ws2)

    # ============================
    # Sheet 3 — Revenue by period
    # ============================
    ws3 = wb.create_sheet("Revenue by Period")
    headers3 = ["Period", "Orders", "Revenue (UZS)", "Avg Order Value (UZS)"]
    for col, h in enumerate(headers3, start=1):
        ws3.cell(row=1, column=col, value=h)
    style_header_row(ws3, 1, len(headers3))

    for row_i, bucket in enumerate(all_buckets, start=2):
        rev = bucket_revenue[bucket]
        cnt = bucket_orders[bucket]
        avg = round(rev / cnt) if cnt else 0
        ws3.cell(row=row_i, column=1, value=bucket)
        ws3.cell(row=row_i, column=2, value=cnt)
        ws3.cell(row=row_i, column=3, value=round(rev))
        ws3.cell(row=row_i, column=4, value=avg)

    auto_width(ws3)

    # ---- Save to buffer ----
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"tejam-report-{start_dt.strftime('%Y%m%d')}-{end_dt.strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
